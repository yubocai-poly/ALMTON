"""
Experiments 1.1 and 1.2

Experiment 1.1: Denser Grids & Success Rates
- 30x30 (900 points) uniform grid
- Metrics: Success rate, median iterations, average SDP solver time
- Visualization: High-resolution "Basin of Attraction" heatmaps

Experiment 1.2: Tabular Testing
- Test 8 algorithms: GD(0.01), GD(0.05), Second-Order Newton, Unregularized Third-Order Newton,
                     AR2-Interp, AR3-Interp, ALMTON, almton_heuristic
- Statistics based only on successfully converged points
- Output to Excel table
"""

import numpy as np
import numpy.linalg as LA
import sys
import time
import pandas as pd
import os
import pickle
from multiprocessing import Pool, cpu_count
from functools import partial
import matplotlib.pyplot as plt
from matplotlib.colors import ListedColormap, BoundaryNorm
from openpyxl import load_workbook
from typing import Optional

DEFAULT_EXPERIMENT_1_RESULTS_DIR = os.path.join(
    os.path.dirname(__file__), "..", "results", "experiment_1"
)
LEGACY_EXPERIMENT_1_RESULTS_DIR = os.path.join(
    os.path.dirname(__file__), "..", "results", "experiment_1_2"
)
DEFAULT_PERF_PROFILE_FUNCTIONS = ["Beale", "Himmelblau", "McCormick", "Bohachevsky"]


def _ensure_dir(path: str) -> None:
    if path:
        os.makedirs(path, exist_ok=True)


def _resolve_data_file(filename: str, preferred_dir: Optional[str] = None) -> str:
    """Return the first existing path for filename across preferred/new/legacy dirs."""
    candidate_dirs = []
    if preferred_dir:
        candidate_dirs.append(preferred_dir)
    candidate_dirs.append(DEFAULT_EXPERIMENT_1_RESULTS_DIR)
    candidate_dirs.append(LEGACY_EXPERIMENT_1_RESULTS_DIR)

    for directory in candidate_dirs:
        if directory is None:
            continue
        candidate_path = os.path.join(directory, filename)
        if os.path.exists(candidate_path):
            return candidate_path
    # Fallback: place under preferred_dir (or default) even if not existing yet
    target_dir = preferred_dir or DEFAULT_EXPERIMENT_1_RESULTS_DIR
    return os.path.join(target_dir, filename)


# Add path
try:
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    src_dir = os.path.join(base_dir, "src")
    if base_dir not in sys.path:
        sys.path.insert(0, base_dir)
    if src_dir not in sys.path:
        sys.path.insert(0, src_dir)
except Exception as e:
    # If path resolution fails, use current working directory
    base_dir = os.getcwd()
    src_dir = os.path.join(base_dir, "src")
    if base_dir not in sys.path:
        sys.path.insert(0, base_dir)
    if src_dir not in sys.path:
        sys.path.insert(0, src_dir)

try:
    from AdaptiveFramework import almton_simple, almton_heuristic
    from AR2 import ar2_interp  # AR2-Simple removed
    from AR3 import ar3_interp
    from NewtonFunctions import init_func, init_params
    from UnregularizedThirdOrder import (
        unregularized_third_newton_run,
        newton_run,
        gradient_descent,
    )
except ImportError as e:
    print(f"Import Error: {e}")
    print(f"Current working directory: {os.getcwd()}")
    print(f"sys.path: {sys.path[:5]}")
    raise


# ============================================================================
# Experiment 1.1: Dense Grid Test
# ============================================================================


def generate_dense_grid(func_name, grid_size=(20, 20)):
    """
    Generate dense grid points.

    Parameters:
    - func_name: Function name
    - grid_size: Grid size (nx, ny)

    Returns:
    - list: List of grid points
    """
    [XMIN, XMAX, YMIN, YMAX, x_min] = init_params(func_name)

    x_vals = np.linspace(XMIN, XMAX, grid_size[0])
    y_vals = np.linspace(YMIN, YMAX, grid_size[1])

    grid_points = []
    for x in x_vals:
        for y in y_vals:
            grid_points.append(np.array([[x], [y]]))

    return grid_points, x_vals, y_vals


def run_single_algorithm_grid(args):
    """
    Run a single algorithm on a single starting point (for parallel computing).

    Parameters:
    - args: (x0, func_name, algorithm_name, max_iterations, tol, param_dict)

    Returns:
    - dict: Result dictionary
    """
    x0, func_name, algorithm_name, max_iterations, tol, param_dict = args

    [fX, fx, dx, d2x, d3x] = init_func(func_name)
    x0 = np.array(x0).reshape(-1, 1)

    result = {
        "algorithm": algorithm_name,
        "x0": x0.flatten(),
        "converged": False,
        "iterations": max_iterations,
        "final_f": np.inf,
        "final_grad_norm": np.inf,
        "sdp_time": 0.0,
        "error": None,
        "failure_reason": None,  # Track why convergence failed
    }

    try:
        start_time = time.time()

        if algorithm_name == "GD (α=0.01)":
            # Gradient descent with backtracking line search, initial step size 0.01
            x_final, converged, iters = gradient_descent(
                fx,
                dx,
                x0,
                alpha_init=0.01,
                max_iter=max_iterations,
                tol=tol,
                use_line_search=True,
            )
            result["iterations"] = iters
            result["converged"] = converged
            if converged:
                result["final_f"] = fx(x_final)
                result["final_grad_norm"] = LA.norm(dx(x_final))
            else:
                if iters >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "gradient_not_sufficiently_small"

        elif algorithm_name == "GD (α=0.05)":
            # Gradient descent with backtracking line search, initial step size 0.05
            x_final, converged, iters = gradient_descent(
                fx,
                dx,
                x0,
                alpha_init=0.05,
                max_iter=max_iterations,
                tol=tol,
                use_line_search=True,
            )
            result["iterations"] = iters
            result["converged"] = converged
            if converged:
                result["final_f"] = fx(x_final)
                result["final_grad_norm"] = LA.norm(dx(x_final))
            else:
                if iters >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "gradient_not_sufficiently_small"

        elif algorithm_name == "Second-Order Newton":
            # Newton's method with line search
            x_final, converged, iters = newton_run(
                fx, dx, d2x, d3x, x0, max_iterations, tol, use_line_search=True
            )

            # Post-processing check: Verify if it converged to a saddle point or local max
            final_val = fx(x_final)
            if converged and final_val > 1.0:
                converged = False
                result["failure_reason"] = "converged_to_high_function_value"

            result["iterations"] = iters
            result["converged"] = converged

            if converged:
                result["final_f"] = fx(x_final)
                result["final_grad_norm"] = LA.norm(dx(x_final))
            else:
                # Check if it's a saddle point based on Hessian
                final_grad_norm = LA.norm(dx(x_final))
                if final_grad_norm <= tol:
                    H_final = d2x(x_final)
                    eigenvalues = LA.eigvalsh(H_final)
                    min_eigenvalue = np.min(eigenvalues)
                    if min_eigenvalue < -1e-6:
                        result["failure_reason"] = "saddle_point"
                    elif final_val > 1.0:
                        result["failure_reason"] = "local_maximum_or_plateau"
                    else:
                        result["failure_reason"] = "unknown"
                elif iters >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "gradient_not_sufficiently_small"
            # Store values even if failed for debugging
            result["final_f"] = final_val
            result["final_grad_norm"] = LA.norm(dx(x_final))

        elif algorithm_name == "Unregularized Third-Order Newton":
            x_final, converged, iters = unregularized_third_newton_run(
                fx, dx, d2x, d3x, x0, max_iterations, tol
            )
            result["iterations"] = iters
            result["converged"] = converged
            if converged:
                result["final_f"] = fx(x_final)
                result["final_grad_norm"] = LA.norm(dx(x_final))
            else:
                if iters >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    # Check if SDP failed
                    result["failure_reason"] = "sdp_subproblem_failure"

        elif algorithm_name == "AR2-Interp":
            ar2_interp_params = param_dict.get(
                "ar2_interp", [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8]
            )
            r = ar2_interp(
                fx, dx, d2x, x0, max_iterations, tol, ar2_interp_params, verbose=False
            )
            result["iterations"] = r["iterations"]
            result["converged"] = r["converged"]
            if r["converged"]:
                result["final_f"] = r["f_history"][-1]
                result["final_grad_norm"] = r["grad_norm_history"][-1]
            else:
                if r["iterations"] >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "subproblem_failure"

        elif algorithm_name == "AR3-Interp":
            ar3_interp_params = param_dict.get(
                "ar3_interp", [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8]
            )
            r = ar3_interp(
                fx,
                dx,
                d2x,
                d3x,
                x0,
                max_iterations,
                tol,
                ar3_interp_params,
                verbose=False,
            )
            result["iterations"] = r["iterations"]
            result["converged"] = r["converged"]
            if r["converged"]:
                result["final_f"] = r["f_history"][-1]
                result["final_grad_norm"] = r["grad_norm_history"][-1]
            else:
                if r["iterations"] >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "subproblem_failure"
            # Record SDP solve time (if available)
            if "sdp_time" in r:
                result["sdp_time"] = r["sdp_time"]

        elif algorithm_name == "ALMTON":
            almton_params = param_dict.get("almton", [0.1, 0.01, 0.1, 2.0])
            r = almton_simple(
                fx, dx, d2x, d3x, x0, max_iterations, tol, almton_params, verbose=False
            )
            result["iterations"] = r["iterations"]
            result["converged"] = r["converged"]
            if r["converged"]:
                result["final_f"] = r["f_history"][-1]
                result["final_grad_norm"] = r["grad_norm_history"][-1]
            else:
                # Determine failure reason
                if r.get("sigma_exceeded", False):
                    result["failure_reason"] = "sigma_exceeded"
                elif r["iterations"] >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "sdp_solver_failure"
            # Record SDP solve time
            if "sdp_time" in r:
                result["sdp_time"] = r["sdp_time"]
            elif "sdp_solves" in r:
                # Estimate SDP time (if solve count is available)
                result["sdp_time"] = r["sdp_solves"] * 0.001  # Assume 1ms per solve

        elif algorithm_name == "almton_heuristic":
            almton_heuristic_params = param_dict.get(
                "almton_heuristic", [0.1, 0.01, 0.1, 2.0]
            )
            r = almton_heuristic(
                fx,
                dx,
                d2x,
                d3x,
                x0,
                max_iterations,
                tol,
                almton_heuristic_params,
                verbose=False,
            )
            result["iterations"] = r["iterations"]
            result["converged"] = r["converged"]
            if r["converged"]:
                result["final_f"] = r["f_history"][-1]
                result["final_grad_norm"] = r["grad_norm_history"][-1]
            else:
                # Determine failure reason
                if r.get("sigma_exceeded", False):
                    result["failure_reason"] = "sigma_exceeded"
                elif r["iterations"] >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "sdp_solver_failure"
            # Record SDP solve time
            if "sdp_time" in r:
                result["sdp_time"] = r["sdp_time"]
            elif "sdp_solves" in r:
                result["sdp_time"] = r["sdp_solves"] * 0.001

        result["total_time"] = time.time() - start_time

    except Exception as e:
        result["error"] = str(e)
        result["converged"] = False
        result["failure_reason"] = "exception"

    return result


def experiment_1_1(
    func_name,
    grid_size=(20, 20),
    max_iterations=100,
    tol=1e-8,
    use_parallel=True,
    n_jobs=None,
    save_path=None,
):
    """
    Experiment 1.1: Dense Grid Test

    Parameters:
    - func_name: Function name
    - grid_size: Grid size
    - max_iterations: Maximum number of iterations
    - tol: Convergence tolerance
    - use_parallel: Whether to use parallel computing
    - n_jobs: Number of parallel processes
    - save_path: Path to save results

    Returns:
    - dict: Result dictionary
    """
    print("=" * 120)
    print(f"Experiment 1.1: Dense Grid Test - {func_name}")
    print("=" * 120)

    if save_path is None:
        save_path = DEFAULT_EXPERIMENT_1_RESULTS_DIR
    _ensure_dir(save_path)

    # Generate grid points
    grid_points, x_vals, y_vals = generate_dense_grid(func_name, grid_size)
    print(f"Generated {len(grid_points)} grid points ({grid_size[0]}x{grid_size[1]})")

    # Algorithm list (AR2-Simple removed)
    algorithms = [
        "GD (α=0.01)",
        "GD (α=0.05)",
        "Second-Order Newton",
        "Unregularized Third-Order Newton",
        "AR2-Interp",
        "AR3-Interp",
        "ALMTON",
        "almton_heuristic",
    ]

    # Parameter settings
    param_dict = {
        "ar2_interp": [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8],
        "ar3_interp": [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8],
        "almton": [0.1, 0.01, 0.1, 2.0],
        "almton_heuristic": [0.1, 0.01, 0.1, 2.0],
    }

    # Determine number of parallel processes
    if n_jobs is None:
        n_jobs = cpu_count()

    # Run all algorithms
    all_results = {}

    for algo_name in algorithms:
        print(f"\nTesting algorithm: {algo_name}")
        print("-" * 120)

        # Prepare arguments
        args_list = [
            (x0, func_name, algo_name, max_iterations, tol, param_dict)
            for x0 in grid_points
        ]

        # Run
        if use_parallel and len(grid_points) > 1:
            with Pool(processes=n_jobs) as pool:
                results = pool.map(run_single_algorithm_grid, args_list)
        else:
            results = []
            for i, args in enumerate(args_list):
                if i % 50 == 0:
                    print(f"  Progress: {i}/{len(args_list)}")
                results.append(run_single_algorithm_grid(args))

        all_results[algo_name] = results
        print(f"  Completed: {algo_name}")

    # Calculate statistics
    print("\n" + "=" * 120)
    print("Statistics")
    print("=" * 120)

    stats = {}
    for algo_name in algorithms:
        results = all_results[algo_name]

        # Only count successfully converged points
        converged_results = [r for r in results if r["converged"]]
        n_converged = len(converged_results)
        n_total = len(results)
        success_rate = n_converged / n_total if n_total > 0 else 0

        if n_converged > 0:
            iterations = [r["iterations"] for r in converged_results]
            median_iters = np.median(iterations)
            q1_iters = np.percentile(iterations, 25)
            q3_iters = np.percentile(iterations, 75)
            iqr_iters = q3_iters - q1_iters

            sdp_times = [r["sdp_time"] for r in converged_results if r["sdp_time"] > 0]
            median_sdp_time = np.median(sdp_times) if len(sdp_times) > 0 else 0.0
        else:
            median_iters = np.nan
            q1_iters = np.nan
            q3_iters = np.nan
            iqr_iters = np.nan
            median_sdp_time = 0.0

        # Track failure reasons
        failure_reasons = {}
        failed_results = [r for r in results if not r["converged"]]
        for r in failed_results:
            reason = r.get("failure_reason", "unknown")
            failure_reasons[reason] = failure_reasons.get(reason, 0) + 1

        stats[algo_name] = {
            "success_rate": success_rate,
            "n_converged": n_converged,
            "n_total": n_total,
            "median_iterations": median_iters,
            "q1_iterations": q1_iters,
            "q3_iterations": q3_iters,
            "iqr_iterations": iqr_iters,
            "median_sdp_time": median_sdp_time,
            "failure_reasons": failure_reasons,
        }

        print(f"\n{algo_name}:")
        print(f"  Success Rate: {success_rate:.1%} ({n_converged}/{n_total})")
        if n_converged > 0:
            print(
                f"  Median Iterations: {median_iters:.1f} (IQR: {q1_iters:.1f}-{q3_iters:.1f})"
            )
            if median_sdp_time > 0:
                print(f"  Median SDP Time: {median_sdp_time:.4f}s")
        if failure_reasons:
            print(f"  Failure Reasons:")
            for reason, count in sorted(failure_reasons.items(), key=lambda x: -x[1]):
                print(f"    {reason}: {count} ({count/len(failed_results)*100:.1f}%)")

    # Save detailed results to pickle file
    if save_path:
        os.makedirs(save_path, exist_ok=True)
        pickle_path = os.path.join(save_path, f"{func_name}_experiment_1_1_data.pkl")
        pickle_data = {
            "all_results": all_results,
            "stats": stats,
            "grid_points": grid_points,
            "x_vals": x_vals,
            "y_vals": y_vals,
            "grid_size": grid_size,
            "func_name": func_name,
            "max_iterations": max_iterations,
            "tol": tol,
        }
        with open(pickle_path, "wb") as f:
            pickle.dump(pickle_data, f)
        print(f"\nAll detailed results saved to: {pickle_path}")

        # Generate heatmap
        plot_convergence_heatmaps_experiment_1_1(
            all_results, func_name, grid_size, x_vals, y_vals, save_path
        )

    return {
        "all_results": all_results,
        "stats": stats,
        "grid_points": grid_points,
        "x_vals": x_vals,
        "y_vals": y_vals,
    }


def plot_convergence_heatmaps_experiment_1_1(
    all_results, func_name, grid_size, x_vals, y_vals, save_path
):
    """
    Generate convergence heatmaps for Experiment 1.1 (2x4 layout: 2 rows, 4 columns, 8 algorithms)
    """
    algorithms = list(all_results.keys())
    nx, ny = grid_size

    # Create 2x4 subplot layout (8 algorithms)
    fig, axes = plt.subplots(2, 4, figsize=(32, 16))
    axes = axes.flatten()

    [XMIN, XMAX, YMIN, YMAX, x_min] = init_params(func_name)

    for idx, algo_name in enumerate(algorithms):
        if idx >= 8:
            break

        ax = axes[idx]
        results = all_results[algo_name]

        # Create grid data
        convergence_grid = np.zeros((nx, ny))
        iteration_grid = np.full((nx, ny), np.nan)

        for i, result in enumerate(results):
            row = i // ny
            col = i % ny
            if result["converged"]:
                convergence_grid[row, col] = 1
                iteration_grid[row, col] = result["iterations"]
            else:
                convergence_grid[row, col] = 0

        # Create heatmap data
        heatmap_data = iteration_grid.copy()
        heatmap_data[np.isnan(heatmap_data)] = -1  # Set unconverged points to -1

        # Create colormap
        colors = ["red"] + plt.cm.PuBu(np.linspace(0, 1, 100)).tolist()
        cmap = ListedColormap(colors)
        vmin, vmax = 0, 100
        bounds = [-1.5, -0.5] + list(np.linspace(vmin, vmax, 100))
        norm = BoundaryNorm(bounds, cmap.N)

        # Draw heatmap
        im = ax.imshow(
            heatmap_data,
            cmap=cmap,
            norm=norm,
            extent=[x_vals.min(), x_vals.max(), y_vals.min(), y_vals.max()],
            origin="lower",
            aspect="auto",
        )

        ax.set_title(f"{algo_name}", fontsize=14, fontweight="bold")
        ax.set_xlabel(r"$x_1$", fontsize=12)
        ax.set_ylabel(r"$x_2$", fontsize=12)
        ax.grid(True, alpha=0.3)

    # Hide extra subplots
    for idx in range(len(algorithms), 8):
        axes[idx].set_visible(False)

    # Add shared colorbar
    cbar_ax = fig.add_axes([0.93, 0.15, 0.015, 0.7])
    colors = ["red"] + plt.cm.PuBu(np.linspace(0, 1, 100)).tolist()
    cmap = ListedColormap(colors)
    bounds = [-1.5, -0.5] + list(np.linspace(0, 100, 100))
    norm = BoundaryNorm(bounds, cmap.N)
    cbar = plt.colorbar(
        plt.cm.ScalarMappable(norm=norm, cmap=cmap),
        cax=cbar_ax,
        boundaries=bounds,
    )
    cbar.set_label("Convergence Iterations", fontsize=12)

    plt.tight_layout(rect=[0, 0, 0.92, 1])

    # Save
    file_path = os.path.join(save_path, f"{func_name}_experiment_1_1_heatmap.pdf")
    plt.savefig(file_path, bbox_inches="tight", dpi=300)
    print(f"\nHeatmap saved to: {file_path}")
    plt.close()


# ============================================================================
# Experiment 1.2: Tabular Test
# ============================================================================


def experiment_1_2(
    func_name,
    n_trials=None,
    max_iterations=100,
    tol=1e-8,
    use_parallel=True,
    n_jobs=None,
    save_path=None,
    use_grid_data=True,
    grid_data_path=None,
):
    """
    Experiment 1.2: Tabular Test

    Parameters:
    - func_name: Function name
    - n_trials: Number of trials (ignored if use_grid_data=True)
    - max_iterations: Maximum number of iterations
    - tol: Convergence tolerance
    - use_parallel: Whether to use parallel computing
    - n_jobs: Number of parallel processes
    - save_path: Path to save results
    - use_grid_data: Whether to use grid data from Experiment 1.1 (default True)
    - grid_data_path: Path to Experiment 1.1 pickle file (if None, auto search)

    Returns:
    - dict: Result dictionary
    """
    print("=" * 120)

    if save_path is None:
        save_path = DEFAULT_EXPERIMENT_1_RESULTS_DIR
    _ensure_dir(save_path)
    print(f"Experiment 1.2: Tabular Test - {func_name}")
    print("=" * 120)

    # Initialize function
    [fX, fx, dx, d2x, d3x] = init_func(func_name)
    [XMIN, XMAX, YMIN, YMAX, x_min] = init_params(func_name)

    # Algorithm list (AR2-Simple removed)
    algorithms = [
        "GD (α=0.01)",
        "GD (α=0.05)",
        "Second-Order Newton",
        "Unregularized Third-Order Newton",
        "AR2-Interp",
        "AR3-Interp",
        "ALMTON",
        "almton_heuristic",
    ]

    # Parameter settings
    param_dict = {
        "ar2_interp": [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8],
        "ar3_interp": [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8],
        "almton": [0.1, 0.01, 0.1, 2.0],
        "almton_heuristic": [0.1, 0.01, 0.1, 2.0],
    }

    # Determine starting points: use grid data from Exp 1.1 or random points
    if use_grid_data:
        # Try to load data from Experiment 1.1
        if grid_data_path is None:
            grid_data_path = _resolve_data_file(
                f"{func_name}_experiment_1_1_data.pkl", preferred_dir=save_path
            )

        if os.path.exists(grid_data_path):
            print(f"\nLoading grid data from Experiment 1.1: {grid_data_path}")
            with open(grid_data_path, "rb") as f:
                grid_data = pickle.load(f)
            start_points = grid_data["grid_points"]
            all_results = grid_data["all_results"]
            print(f"  Loaded data for {len(start_points)} grid points")
            print(f"  Using full results from Experiment 1.1 for analysis")
        else:
            print(f"\nWarning: Experiment 1.1 data file not found: {grid_data_path}")
            print(f"  Will use random initial points (n_trials={n_trials or 10})")
            use_grid_data = False

    if not use_grid_data:
        # Generate random initial points (fallback)
        n_trials = n_trials or 10
        print(f"\nGenerating {n_trials} random initial points...")
        start_points = []
        for _ in range(n_trials):
            x0 = np.array(
                [[np.random.uniform(XMIN, XMAX)], [np.random.uniform(YMIN, YMAX)]]
            )
            start_points.append(x0)
        all_results = None

    # If using Experiment 1.1 data, use existing results; otherwise run algorithms
    if all_results is None:
        # Determine number of parallel processes
        if n_jobs is None:
            n_jobs = cpu_count()

        # Run all algorithms
        all_results = {}

        for algo_name in algorithms:
            print(f"\nTesting algorithm: {algo_name}")

            # Prepare arguments
            args_list = [
                (x0, func_name, algo_name, max_iterations, tol, param_dict)
                for x0 in start_points
            ]

            # Run
            if use_parallel and len(start_points) > 1:
                with Pool(processes=n_jobs) as pool:
                    results = pool.map(run_single_algorithm_grid, args_list)
            else:
                results = []
                for args in args_list:
                    results.append(run_single_algorithm_grid(args))

            all_results[algo_name] = results
            print(f"  Completed: {algo_name}")
    else:
        print(
            f"\nUsing existing results from Experiment 1.1, skipping algorithm execution"
        )

    # Calculate statistics (only for successfully converged points)
    print("\n" + "=" * 120)
    print("Statistics (Only for successfully converged points)")
    print("=" * 120)

    stats = {}
    detailed_data = []

    for algo_name in algorithms:
        # Skip AR2-Simple if it exists in loaded results but we removed it from algorithms list
        if algo_name not in all_results:
            continue

        results = all_results[algo_name]

        # Only count successfully converged points
        converged_results = [r for r in results if r["converged"]]
        n_converged = len(converged_results)
        n_total = len(results)
        conv_rate = n_converged / n_total if n_total > 0 else 0

        if n_converged > 0:
            iterations = [r["iterations"] for r in converged_results]
            final_f = [r["final_f"] for r in converged_results]
            times = [r["total_time"] for r in converged_results]

            median_iters = np.median(iterations)
            q1_iters = np.percentile(iterations, 25)
            q3_iters = np.percentile(iterations, 75)
            iqr_iters = q3_iters - q1_iters

            median_time = np.median(times)
            q1_time = np.percentile(times, 25)
            q3_time = np.percentile(times, 75)
            iqr_time = q3_time - q1_time

            median_final_f = np.median(final_f)
        else:
            median_iters = np.nan
            q1_iters = np.nan
            q3_iters = np.nan
            iqr_iters = np.nan
            median_time = np.nan
            q1_time = np.nan
            q3_time = np.nan
            iqr_time = np.nan
            median_final_f = np.nan

        # Statistics on failure reasons
        failure_reasons = {}
        failed_results = [r for r in results if not r["converged"]]
        for r in failed_results:
            reason = r.get("failure_reason", "unknown")
            failure_reasons[reason] = failure_reasons.get(reason, 0) + 1

        stats[algo_name] = {
            "median_iters": median_iters,
            "q1_iters": q1_iters,
            "q3_iters": q3_iters,
            "iqr_iters": iqr_iters,
            "median_time": median_time,
            "q1_time": q1_time,
            "q3_time": q3_time,
            "iqr_time": iqr_time,
            "conv_rate": conv_rate,
            "median_final_f": median_final_f,
            "n_converged": n_converged,
            "n_total": n_total,
            "failure_reasons": failure_reasons,
        }

        # Detailed data
        for i, result in enumerate(results):
            detailed_data.append(
                {
                    "Trial": i + 1,
                    "Algorithm": algo_name,
                    "Initial x0": str(result["x0"]),
                    "Converged": result["converged"],
                    "Iterations": (
                        result["iterations"] if result["converged"] else np.nan
                    ),
                    "Final f(x)": result["final_f"] if result["converged"] else np.nan,
                    "Final ||∇f||": (
                        result["final_grad_norm"] if result["converged"] else np.nan
                    ),
                    "Time (s)": result["total_time"],
                    "Failure Reason": result.get("failure_reason", ""),
                    "Error": result["error"] if result["error"] else "",
                }
            )

    # Print statistics table
    print(
        f"\n{'Algorithm':<30} {'Median Iters (IQR)':<25} {'Median Time(s) (IQR)':<25} {'Conv Rate':<12} {'Median Final f':<15}"
    )
    print("-" * 120)

    for algo_name in algorithms:
        if algo_name not in stats:
            continue
        s = stats[algo_name]
        if not np.isnan(s["median_iters"]):
            iters_str = (
                f"{s['median_iters']:.2f} ({s['q1_iters']:.2f}-{s['q3_iters']:.2f})"
            )
            time_str = f"{s['median_time']:.4f} ({s['q1_time']:.4f}-{s['q3_time']:.4f})"
            print(
                f"{algo_name:<30} {iters_str:<25} {time_str:<25} "
                f"{s['conv_rate']:>6.1%}       {s['median_final_f']:>12.6e}"
            )
        else:
            print(
                f"{algo_name:<30} {'N/A':<25} {'N/A':<25} {s['conv_rate']:>6.1%}       {'N/A':<15}"
            )
        # Print failure reason statistics
        if s["failure_reasons"]:
            print(f"  Failure Reasons:")
            for reason, count in sorted(
                s["failure_reasons"].items(), key=lambda x: -x[1]
            ):
                print(
                    f"    {reason}: {count} ({count/(s['n_total']-s['n_converged'])*100:.1f}%)"
                )

    # Save to Excel
    if save_path:
        os.makedirs(save_path, exist_ok=True)

        # Create DataFrame
        df_detailed = pd.DataFrame(detailed_data)

        # Create summary table, including failure reason statistics
        summary_rows = []
        for algo_name in algorithms:
            if algo_name not in stats:
                continue
            s = stats[algo_name]
            row = {
                "Algorithm": algo_name,
                "Median Iters": s["median_iters"],
                "Q1 Iters": s["q1_iters"],
                "Q3 Iters": s["q3_iters"],
                "IQR Iters": s["iqr_iters"],
                "Median Time(s)": s["median_time"],
                "Q1 Time(s)": s["q1_time"],
                "Q3 Time(s)": s["q3_time"],
                "IQR Time(s)": s["iqr_time"],
                "Conv Rate": s["conv_rate"],
                "Median Final f": s["median_final_f"],
                "N Converged": s["n_converged"],
                "N Total": s["n_total"],
            }
            # Add failure reason statistics (as string)
            if s["failure_reasons"]:
                failure_str = "; ".join(
                    [
                        f"{k}:{v}"
                        for k, v in sorted(
                            s["failure_reasons"].items(), key=lambda x: -x[1]
                        )
                    ]
                )
                row["Failure Reasons"] = failure_str
            else:
                row["Failure Reasons"] = ""
            summary_rows.append(row)

        df_summary = pd.DataFrame(summary_rows)

        # Format numeric columns to 6 decimal places
        # Summary table numeric columns
        numeric_cols_summary = [
            "Median Iters",
            "Q1 Iters",
            "Q3 Iters",
            "IQR Iters",
            "Median Time(s)",
            "Q1 Time(s)",
            "Q3 Time(s)",
            "IQR Time(s)",
            "Conv Rate",
            "Median Final f",
        ]
        for col in numeric_cols_summary:
            if col in df_summary.columns:
                df_summary[col] = pd.to_numeric(df_summary[col], errors="coerce").round(
                    6
                )

        # Detailed table numeric columns
        numeric_cols_detailed = ["Iterations", "Final f(x)", "Final ||∇f||", "Time (s)"]
        for col in numeric_cols_detailed:
            if col in df_detailed.columns:
                df_detailed[col] = pd.to_numeric(
                    df_detailed[col], errors="coerce"
                ).round(6)

        # Save
        excel_path = os.path.join(save_path, f"{func_name}_experiment_1_2.xlsx")
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="Summary", index=False)
            df_detailed.to_excel(writer, sheet_name="Detailed", index=False)

        # Use openpyxl to format cells, ensure 6 decimal places display
        try:
            wb = load_workbook(excel_path)

            # Format Summary sheet
            ws_summary = wb["Summary"]
            # Find numeric column indices
            header_row = 1
            col_mapping_summary = {}
            for col_idx, cell in enumerate(ws_summary[header_row], start=1):
                if cell.value in numeric_cols_summary:
                    col_mapping_summary[cell.value] = col_idx

            # Apply formatting
            for row_idx in range(2, ws_summary.max_row + 1):
                for col_name, col_idx in col_mapping_summary.items():
                    cell = ws_summary.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        try:
                            float_val = float(cell.value)
                            if not np.isnan(float_val):
                                cell.number_format = "0.000000"
                        except (ValueError, TypeError):
                            pass

            # Format Detailed sheet
            ws_detailed = wb["Detailed"]
            # Find numeric column indices
            col_mapping_detailed = {}
            for col_idx, cell in enumerate(ws_detailed[header_row], start=1):
                if cell.value in numeric_cols_detailed:
                    col_mapping_detailed[cell.value] = col_idx

            # Apply formatting
            for row_idx in range(2, ws_detailed.max_row + 1):
                for col_name, col_idx in col_mapping_detailed.items():
                    cell = ws_detailed.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        try:
                            float_val = float(cell.value)
                            if not np.isnan(float_val):
                                cell.number_format = "0.000000"
                        except (ValueError, TypeError):
                            pass

            wb.save(excel_path)
        except Exception as e:
            print(f"Warning: Error formatting Excel cells: {e}")
            print(f"Data saved, but 6 decimal formatting may not be applied")

        print(f"\nResults saved to: {excel_path}")

        # Also save pickle file
        pickle_path = os.path.join(save_path, f"{func_name}_experiment_1_2_data.pkl")
        pickle_data = {
            "all_results": all_results,
            "stats": stats,
            "detailed_data": detailed_data,
            "func_name": func_name,
            "max_iterations": max_iterations,
            "tol": tol,
            "use_grid_data": use_grid_data,
        }
        with open(pickle_path, "wb") as f:
            pickle.dump(pickle_data, f)
        print(f"Detailed data saved to: {pickle_path}")

    return {
        "all_results": all_results,
        "stats": stats,
        "detailed_data": detailed_data,
    }


# ============================================================================
# Experiment 1.3: Dolan-Moré Performance Profile
# ============================================================================


def collect_profile_data_experiment_13(
    test_functions,
    save_path,
    grid_size=(30, 30),
    max_iterations=100,
    tol=1e-8,
    regenerate=False,
):
    """
    Ensure Experiment 1.1 data exists for the requested functions and load it.
    """
    _ensure_dir(save_path)
    all_data = {}

    for func_name in test_functions:
        print("\n" + "=" * 80)
        print(f"Preparing data for Experiment 1.3 - {func_name}")
        print("=" * 80)

        pickle_name = f"{func_name}_experiment_1_1_data.pkl"
        pickle_path = os.path.join(save_path, pickle_name)

        if regenerate or not os.path.exists(pickle_path):
            resolved_path = _resolve_data_file(pickle_name, preferred_dir=save_path)
            if not regenerate and os.path.exists(resolved_path):
                pickle_path = resolved_path
            else:
                print(
                    f"  Data not found or regeneration requested. Running Experiment 1.1 ({grid_size[0]}x{grid_size[1]} grid)..."
                )
                experiment_1_1(
                    func_name=func_name,
                    grid_size=grid_size,
                    max_iterations=max_iterations,
                    tol=tol,
                    use_parallel=True,
                    save_path=save_path,
                )
                pickle_path = os.path.join(save_path, pickle_name)

        print(f"  Loading data from: {pickle_path}")
        with open(pickle_path, "rb") as f:
            func_data = pickle.load(f)
        all_data[func_name] = func_data

    return all_data


def compute_performance_ratios_experiment_13(all_data, metric="iterations"):
    """
    Compute performance ratios for Dolan-Moré profile.
    """
    data_list = []

    for func_name, func_data in all_data.items():
        all_results = func_data["all_results"]
        grid_points = func_data["grid_points"]
        algorithms = list(all_results.keys())

        num_points = len(grid_points)
        for point_idx in range(num_points):
            for algo_name in algorithms:
                if point_idx >= len(all_results[algo_name]):
                    continue
                run_result = all_results[algo_name][point_idx]

                if run_result["converged"]:
                    if metric == "iterations":
                        value = run_result["iterations"]
                    elif metric == "total_time":
                        value = run_result.get("total_time", np.inf)
                    else:
                        raise ValueError(f"Unknown metric: {metric}")
                    if value <= 0:
                        value = 1e-8
                else:
                    value = np.inf

                data_list.append(
                    {
                        "problem_id": f"{func_name}_{point_idx}",
                        "algorithm": algo_name,
                        "value": value,
                        "converged": run_result["converged"],
                    }
                )

    df = pd.DataFrame(data_list)
    pivot_df = df.pivot(index="problem_id", columns="algorithm", values="value")
    min_vals = pivot_df.min(axis=1)
    valid_problems = min_vals < np.inf

    if not valid_problems.all():
        print(
            f"Warning: {np.sum(~valid_problems)} problems had no successful solver across algorithms."
        )

    ratios = pivot_df.divide(min_vals, axis=0)
    ratios = ratios.replace([np.inf, -np.inf], np.inf)
    ratios = ratios.fillna(np.inf)

    metadata = {
        "problem_ids": pivot_df.index.tolist(),
        "algorithms": pivot_df.columns.tolist(),
        "n_problems": len(pivot_df),
        "n_algorithms": len(pivot_df.columns),
        "valid_problems": int(valid_problems.sum()),
    }

    return ratios, metadata


def compute_performance_profile_experiment_13(ratios, tau_max=100, n_points=1000):
    tau_values = np.logspace(0, np.log10(tau_max), n_points)
    n_problems = len(ratios)

    profiles = {}
    for algo in ratios.columns:
        algo_ratios = ratios[algo].values
        sorted_ratios = np.sort(algo_ratios)
        rho_values = np.zeros_like(tau_values)

        for i, tau in enumerate(tau_values):
            count = np.sum(sorted_ratios <= tau)
            rho_values[i] = count / n_problems

        profiles[algo] = (tau_values, rho_values)

    return profiles


def plot_dolan_more_profile_experiment_13(
    profiles,
    metadata,
    metric="iterations",
    save_path=None,
    tau_max=100,
    figsize=(10, 6),
):
    plt.figure(figsize=figsize)

    # Display names for legend (data keys may still be ALMTON / almton_heuristic)
    legend_display_names = {
        "ALMTON": "ALMTON-Simple",
        "almton_heuristic": "ALMTON-Heuristic",
        "ALMTON-Interp": "ALMTON-Interp",
    }
    colors = {
        "GD (α=0.01)": "blue",
        "GD (α=0.05)": "lightblue",
        "Second-Order Newton": "orange",
        "Unregularized Third-Order Newton": "purple",
        "AR2-Interp": "green",
        "AR3-Interp": "darkgreen",
        "ALMTON": "red",
        "almton_heuristic": "darkred",
        "ALMTON-Interp": "crimson",
    }

    for algo, (tau_vals, rho_vals) in profiles.items():
        color = colors.get(algo, "black")
        label = legend_display_names.get(algo, algo)
        plt.plot(
            tau_vals,
            rho_vals,
            label=label,
            linewidth=2,
            color=color,
            alpha=0.85,
        )

    plt.xscale("log")
    plt.xlim(1, tau_max)
    plt.ylim(0, 1.05)
    plt.xlabel(r"Performance Ratio $\tau$ (Log Scale)", fontsize=12)
    plt.ylabel(r"Probability $\rho_s(\tau)$", fontsize=12)
    plt.title(
        f"Experiment 1.3 Performance Profile ({metric.capitalize()})\n"
        f"{metadata['n_problems']} problems, {metadata['valid_problems']} with at least one successful solver",
        fontsize=14,
    )
    plt.legend(loc="lower right", fontsize=10)
    plt.grid(True, which="both", ls="-", alpha=0.2)
    plt.tight_layout()

    if save_path:
        _ensure_dir(save_path)
        filename = f"experiment_1_3_profile_{metric}.pdf"
        filepath = os.path.join(save_path, filename)
        plt.savefig(filepath, bbox_inches="tight", dpi=300)
        print(f"\nPerformance profile saved to: {filepath}")
    else:
        plt.show()

    plt.close()


def print_profile_statistics_experiment_13(profiles, metadata, ratios):
    print("\n" + "=" * 80)
    print("Experiment 1.3 Performance Profile Statistics")
    print("=" * 80)
    print(f"\nTotal problems: {metadata['n_problems']}")
    print(f"Problems with at least one successful solver: {metadata['valid_problems']}")
    print(
        f"\n{'Algorithm':<30} {'Efficiency (τ=1)':<20} {'Robustness (τ→∞)':<20} {'Successful Cases':<18}"
    )
    print("-" * 100)

    for algo, (tau_vals, rho_vals) in profiles.items():
        efficiency_idx = np.argmin(np.abs(tau_vals - 1.0))
        efficiency = rho_vals[efficiency_idx]
        robustness = rho_vals[-1]
        algo_ratios = ratios[algo].values
        n_success = np.sum(algo_ratios < np.inf)

        print(
            f"{algo:<30} {efficiency:>6.2%} {'':<8} {robustness:>6.2%} ({n_success}/{metadata['n_problems']})"
        )


def experiment_1_3(
    test_functions=None,
    grid_size=(30, 30),
    max_iterations=100,
    tol=1e-8,
    metrics=None,
    save_path=None,
    tau_max=100,
    regenerate=False,
):
    """
    Experiment 1.3: Dolan-Moré Performance Profile on Experiment 1 grid data.
    """
    print("=" * 80)
    print("Experiment 1.3: Dolan-Moré Performance Profile")
    print("=" * 80)

    if save_path is None:
        save_path = DEFAULT_EXPERIMENT_1_RESULTS_DIR
    _ensure_dir(save_path)

    if test_functions is None:
        test_functions = DEFAULT_PERF_PROFILE_FUNCTIONS

    if metrics is None:
        metrics = ["iterations", "total_time"]

    all_data = collect_profile_data_experiment_13(
        test_functions=test_functions,
        save_path=save_path,
        grid_size=grid_size,
        max_iterations=max_iterations,
        tol=tol,
        regenerate=regenerate,
    )

    results = {}
    for metric in metrics:
        print("\n" + "=" * 80)
        print(f"Computing performance profile for metric: {metric}")
        print("=" * 80)

        ratios, metadata = compute_performance_ratios_experiment_13(
            all_data, metric=metric
        )
        profiles = compute_performance_profile_experiment_13(ratios, tau_max=tau_max)
        print_profile_statistics_experiment_13(profiles, metadata, ratios)
        plot_dolan_more_profile_experiment_13(
            profiles,
            metadata,
            metric=metric,
            save_path=save_path,
            tau_max=tau_max,
        )

        results[metric] = {
            "ratios": ratios,
            "profiles": profiles,
            "metadata": metadata,
        }

    pickle_path = os.path.join(save_path, "experiment_1_3_results.pkl")
    with open(pickle_path, "wb") as f:
        pickle.dump(
            {
                "all_data": all_data,
                "results": results,
                "test_functions": test_functions,
                "grid_size": grid_size,
                "max_iterations": max_iterations,
                "tol": tol,
                "tau_max": tau_max,
            },
            f,
        )
    print(f"\nAll Experiment 1.3 results saved to: {pickle_path}")

    return results


# ============================================================================
# Main Function
# ============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Run Experiment 1.1 and 1.2")
    parser.add_argument(
        "--function",
        type=str,
        default="Himmelblau",
        help="Test function name",
    )
    parser.add_argument(
        "--experiment",
        type=str,
        default="1.2",
        choices=["1.1", "1.2", "1.3", "both", "all"],
        help="Experiment type: 1.1, 1.2, 1.3, both (1.1+1.2), or all (1.1+1.2+1.3)",
    )
    parser.add_argument(
        "--trials",
        type=int,
        default=10,
        help="Number of trials for Experiment 1.2",
    )
    parser.add_argument(
        "--grid_size",
        type=int,
        nargs=2,
        default=[20, 20],
        help="Grid size for Experiment 1.1",
    )
    parser.add_argument(
        "--max_iterations",
        type=int,
        default=100,
        help="Maximum iterations",
    )
    parser.add_argument(
        "--tol",
        type=float,
        default=1e-8,
        help="Convergence tolerance",
    )
    parser.add_argument(
        "--no_parallel",
        action="store_true",
        help="Disable parallel computing",
    )
    parser.add_argument(
        "--n_jobs",
        type=int,
        default=None,
        help="Number of parallel processes",
    )
    parser.add_argument(
        "--save_path",
        type=str,
        default=None,
        help="Save path (default: results/experiment_1/)",
    )
    parser.add_argument(
        "--functions_13",
        type=str,
        nargs="+",
        default=None,
        help="Function names for Experiment 1.3 (default: Beale Himmelblau McCormick Bohachevsky)",
    )
    parser.add_argument(
        "--grid_size_13",
        type=int,
        nargs=2,
        default=[30, 30],
        help="Grid size for Experiment 1.3 data regeneration (default: 30 30)",
    )
    parser.add_argument(
        "--metrics_13",
        type=str,
        nargs="+",
        default=["iterations", "total_time"],
        choices=["iterations", "total_time"],
        help="Metrics for Experiment 1.3 performance profiles (default: iterations total_time)",
    )
    parser.add_argument(
        "--tau_max_13",
        type=float,
        default=100,
        help="Maximum tau value for Experiment 1.3 performance profiles (default: 100)",
    )
    parser.add_argument(
        "--regenerate_13",
        action="store_true",
        help="Regenerate Experiment 1.1 data before running Experiment 1.3",
    )

    args = parser.parse_args()

    # Set save path
    if args.save_path is None:
        args.save_path = DEFAULT_EXPERIMENT_1_RESULTS_DIR

    requested_function = args.function.strip()
    run_all_functions = requested_function.lower() == "all"
    functions_to_run = (
        DEFAULT_PERF_PROFILE_FUNCTIONS.copy()
        if run_all_functions
        else [requested_function]
    )
    grid_size_tuple = tuple(args.grid_size)

    # Run experiment
    if args.experiment in ["1.1", "both", "all"]:
        for func in functions_to_run:
            print("\n" + "=" * 120)
            print(f"Starting Experiment 1.1 - {func}")
            print("=" * 120)
            experiment_1_1(
                func,
                grid_size=grid_size_tuple,
                max_iterations=args.max_iterations,
                tol=args.tol,
                use_parallel=not args.no_parallel,
                n_jobs=args.n_jobs,
                save_path=args.save_path,
            )

    if args.experiment in ["1.2", "both", "all"]:
        for func in functions_to_run:
            print("\n" + "=" * 120)
            print(f"Starting Experiment 1.2 - {func}")
            print("=" * 120)
            experiment_1_2(
                func,
                n_trials=args.trials,
                max_iterations=args.max_iterations,
                tol=args.tol,
                use_parallel=not args.no_parallel,
                n_jobs=args.n_jobs,
                save_path=args.save_path,
                use_grid_data=True,
            )

    if args.experiment in ["1.3", "all"]:
        print("\n" + "=" * 120)
        print("Starting Experiment 1.3")
        print("=" * 120)
        if args.functions_13:
            perf_functions = args.functions_13
        elif run_all_functions:
            perf_functions = functions_to_run
        else:
            perf_functions = DEFAULT_PERF_PROFILE_FUNCTIONS

        experiment_1_3(
            test_functions=perf_functions,
            grid_size=tuple(args.grid_size_13),
            max_iterations=args.max_iterations,
            tol=args.tol,
            metrics=args.metrics_13,
            save_path=args.save_path,
            tau_max=args.tau_max_13,
            regenerate=args.regenerate_13,
        )

    print("\n" + "=" * 120)
    print("All experiments completed!")
    print("=" * 120)
