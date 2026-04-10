"""
Experiment 3: High-Order Geometric Structure Test Set

Experiment 3.1: Trajectory Comparison on High-Order Geometric Structures
- Purpose: Compare ALMTON (third-order) with AR2/Newton (second-order) on functions
  designed to "trap" second-order methods. These functions have significant third-order
  curvature that only third-order methods can "see" to find the correct descent path.
- Functions: Slalom Function, Hairpin Turn Function (from Cartis AR3 paper)
- Expected Results: AR2/Newton may follow zig-zagging paths, ALMTON should find "shortcuts"

Experiment 3.2: Dense Grid & Tabular Analysis (Slalom & HairpinTurn)
- Experiment 3.2.1: Dense Grid Test (20x20 grid, heatmaps)
- Experiment 3.2.2: Tabular Testing (statistical comparison)

Reference: Cartis et al. AR3 paper, Figure 1
"""

import numpy as np
import numpy.linalg as LA
import sys
import time
import os
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.colors import LinearSegmentedColormap, ListedColormap, BoundaryNorm
import pickle
from multiprocessing import Pool, cpu_count
from functools import partial
from openpyxl import load_workbook

# Add path
try:
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    src_dir = os.path.join(base_dir, "src")
    if base_dir not in sys.path:
        sys.path.insert(0, base_dir)
    if src_dir not in sys.path:
        sys.path.insert(0, src_dir)
except Exception as e:
    base_dir = os.getcwd()
    src_dir = os.path.join(base_dir, "src")
    if base_dir not in sys.path:
        sys.path.insert(0, base_dir)
    if src_dir not in sys.path:
        sys.path.insert(0, src_dir)

try:
    from AdaptiveFramework import almton_simple, almton_heuristic, almton_interp
    from AR2 import ar2_interp
    from AR3 import ar3_interp
    from UnregularizedThirdOrder import (
        newton_run,
        gradient_descent,
        unregularized_third_newton_run,
    )
    from NewtonFunctions import init_func, init_params
except ImportError as e:
    print(f"Import Error: {e}")
    print(f"Current working directory: {os.getcwd()}")
    print(f"sys.path: {sys.path[:5]}")
    raise

EXPERIMENT_3_RESULTS_DIR = os.path.join(
    os.path.dirname(__file__), "..", "results", "experiment_3"
)


def _ensure_dir(path: str) -> None:
    if path:
        os.makedirs(path, exist_ok=True)


# ============================================================================
# Algorithm Wrappers with Path Tracking
# ============================================================================


def run_almton_with_path(
    fx, dx, d2x, d3x, x0, max_iterations, tol, param_list, sdp_tol=1e-3
):
    """Run ALMTON and track iteration path"""
    result = almton_simple(
        fx,
        dx,
        d2x,
        d3x,
        x0,
        max_iterations,
        tol,
        param_list,
        verbose=False,
        sdp_tol=sdp_tol,
    )
    # Extract path from history
    path = np.array([x.flatten() for x in result["x_history"]])
    return result, path


def run_almton_heuristic_with_path(
    fx, dx, d2x, d3x, x0, max_iterations, tol, param_list, sdp_tol=1e-3
):
    """Run almton_heuristic and track iteration path"""
    result = almton_heuristic(
        fx,
        dx,
        d2x,
        d3x,
        x0,
        max_iterations,
        tol,
        param_list,
        verbose=False,
        sdp_tol=sdp_tol,
    )
    # Extract path from history
    path = np.array([x.flatten() for x in result["x_history"]])
    return result, path


def run_ar2_with_path(fx, dx, d2x, x0, max_iterations, tol, param_list):
    """Run AR2-Interp and track iteration path"""

    # AR2 doesn't need d3x
    def d3x_dummy(x):
        return np.zeros((2, 2, 2))

    result = ar2_interp(fx, dx, d2x, x0, max_iterations, tol, param_list, verbose=False)
    # Extract path from history
    path = np.array([x.flatten() for x in result["x_history"]])
    return result, path


def run_ar3_with_path(fx, dx, d2x, d3x, x0, max_iterations, tol, param_list):
    """Run AR3-Interp and track iteration path"""
    result = ar3_interp(
        fx, dx, d2x, d3x, x0, max_iterations, tol, param_list, verbose=False
    )
    # Extract path from history
    path = np.array([x.flatten() for x in result["x_history"]])
    return result, path


def run_almton_interp_with_path(
    fx, dx, d2x, d3x, x0, max_iterations, tol, param_list, sdp_tol=1e-8
):
    """Run ALMTON-Interp and track iteration path"""
    result = almton_interp(
        fx,
        dx,
        d2x,
        d3x,
        x0,
        max_iterations,
        tol,
        param_list,
        verbose=False,
        sdp_solver="auto",
    )
    # Extract path from history
    if "x_history" in result:
        path = np.array([x.flatten() for x in result["x_history"]])
    else:
        # Fallback: create simple path
        path = np.array([x0.flatten(), result["x_final"].flatten()])
    return result, path


def run_newton_with_path(fx, dx, d2x, d3x, x0, max_iterations, tol):
    """Run Newton method and track iteration path manually"""
    # Manually track path since newton_run doesn't return history
    x_k = x0.copy()
    path = [x_k.flatten()]
    k = 0
    converged = False

    try:
        while k < max_iterations:
            grad = dx(x_k)
            grad_norm = LA.norm(grad)
            if grad_norm <= tol:
                converged = True
                break

            hessian = d2x(x_k)
            try:
                L = LA.cholesky(hessian)
                p = -LA.solve(L @ L.T, grad)
            except LA.LinAlgError:
                # Use pseudo-inverse if not positive definite
                p = -LA.lstsq(hessian, grad, rcond=None)[0]
                p = p.reshape(-1, 1)

            # Simple line search (Armijo)
            alpha = 1.0
            c = 1e-4
            rho = 0.5
            f_x = fx(x_k)
            grad_dot_p = grad.flatten() @ p.flatten()

            for _ in range(20):
                x_new = x_k + alpha * p
                f_new = fx(x_new)
                if f_new <= f_x + c * alpha * grad_dot_p:
                    break
                alpha = rho * alpha

            x_k = x_k + alpha * p
            path.append(x_k.flatten())
            k += 1

    except Exception as e:
        print(f"Newton method error: {e}")

    path = np.array(path)
    result = {
        "x_final": x_k,
        "converged": converged,
        "iterations": k,
    }
    return result, path


# ============================================================================
# Visualization Functions
# ============================================================================


def plot_trajectory_comparison(
    func_name, paths_dict, x0, x_min, XMIN, XMAX, YMIN, YMAX, save_path=None
):
    """
    Plot trajectory comparison for different algorithms.

    Parameters:
    - func_name: Name of the function
    - paths_dict: Dictionary of {algorithm_name: path_array}
    - x0: Starting point
    - x_min: Minimum point(s)
    - XMIN, XMAX, YMIN, YMAX: Plotting bounds
    - save_path: Path to save figure
    """
    # Create figure with subplots
    fig, axes = plt.subplots(1, 2, figsize=(20, 10))

    # Generate contour plot data
    x_vals = np.linspace(XMIN, XMAX, 200)
    y_vals = np.linspace(YMIN, YMAX, 200)
    X_grid, Y_grid = np.meshgrid(x_vals, y_vals)

    [fX, fx, dx, d2x, d3x] = init_func(func_name)
    Z_grid = np.zeros_like(X_grid)
    for i in range(X_grid.shape[0]):
        for j in range(X_grid.shape[1]):
            Z_grid[i, j] = fx(np.array([[X_grid[i, j]], [Y_grid[i, j]]]))

    # Color scheme for algorithms (data keys)
    colors = {
        "ALMTON": "red",
        "almton_heuristic": "darkred",
        "ALMTON-Interp": "crimson",
        "AR2-Interp": "blue",
        "AR3-Interp": "green",
        "Newton": "orange",
    }
    markers = {
        "ALMTON": "o",
        "almton_heuristic": "s",
        "ALMTON-Interp": "p",
        "AR2-Interp": "^",
        "AR3-Interp": "v",
        "Newton": "D",
    }
    # Legend display names for paper
    legend_display_names = {
        "ALMTON": "ALMTON-Simple",
        "almton_heuristic": "ALMTON-Heuristic",
        "ALMTON-Interp": "ALMTON-Interp",
    }

    # Plot 1: Contour plot with trajectories
    ax1 = axes[0]
    contour = ax1.contour(X_grid, Y_grid, Z_grid, levels=30, alpha=0.6, cmap="viridis")
    ax1.clabel(contour, inline=True, fontsize=8)

    # Plot trajectories
    for algo_name, path in paths_dict.items():
        if path is not None and len(path) > 0:
            color = colors.get(algo_name, "black")
            marker = markers.get(algo_name, "o")
            label = legend_display_names.get(algo_name, algo_name)
            ax1.plot(
                path[:, 0],
                path[:, 1],
                color=color,
                marker=marker,
                markersize=6,
                linewidth=2,
                label=label,
                alpha=0.8,
            )
            # Mark start point
            if len(path) > 0:
                ax1.plot(
                    path[0, 0],
                    path[0, 1],
                    color=color,
                    marker="*",
                    markersize=15,
                    markeredgecolor="black",
                    markeredgewidth=1,
                )

    # Mark minimum point
    if x_min is not None and len(x_min) > 0:
        for xm in x_min:
            ax1.plot(xm[0], xm[1], "k*", markersize=20, label="Minimum")

    ax1.set_xlabel("x", fontsize=12)
    ax1.set_ylabel("y", fontsize=12)
    ax1.set_title(
        f"{func_name} - Trajectory Comparison", fontsize=14, fontweight="bold"
    )
    ax1.legend(loc="best", fontsize=10)
    ax1.grid(True, alpha=0.3)
    ax1.set_xlim(XMIN, XMAX)
    ax1.set_ylim(YMIN, YMAX)

    # Plot 2: Function value along trajectory
    ax2 = axes[1]
    for algo_name, path in paths_dict.items():
        if path is not None and len(path) > 0:
            color = colors.get(algo_name, "black")
            marker = markers.get(algo_name, "o")
            label = legend_display_names.get(algo_name, algo_name)
            f_vals = [fx(p.reshape(-1, 1)) for p in path]
            iterations = np.arange(len(f_vals))
            ax2.plot(
                iterations,
                f_vals,
                color=color,
                marker=marker,
                markersize=6,
                linewidth=2,
                label=label,
                alpha=0.8,
            )

    ax2.set_xlabel("Iteration", fontsize=12)
    ax2.set_ylabel("Function Value", fontsize=12)
    ax2.set_title(f"{func_name} - Convergence", fontsize=14, fontweight="bold")
    ax2.legend(loc="best", fontsize=10)
    ax2.grid(True, alpha=0.3)
    ax2.set_yscale("log")

    plt.tight_layout()

    if save_path:
        os.makedirs(os.path.dirname(save_path), exist_ok=True)
        plt.savefig(save_path, dpi=300, bbox_inches="tight")
        print(f"Figure saved to: {save_path}")

    return fig


# ============================================================================
# Main Experiment Function
# ============================================================================


def experiment_3_trajectory_comparison(
    func_name="Slalom",
    x0=None,
    max_iterations=100,
    tol=1e-6,
    save_path=None,
    sdp_tol=1e-3,
):
    """
    Experiment 3.1: Trajectory Comparison on High-Order Geometric Structures

    Parameters:
    - func_name: Function name ('Slalom' or 'HairpinTurn')
    - x0: Starting point (default: from init_params)
    - max_iterations: Maximum iterations (default: 100)
    - tol: Convergence tolerance (default: 1e-6)
    - save_path: Path to save results (default: results/experiment_3/)
    - sdp_tol: SDP tolerance for ALMTON (default: 1e-3)

    Returns:
    - dict: Results dictionary
    """
    print("=" * 120)
    print(f"Experiment 3.1: Trajectory Comparison - {func_name}")
    print("=" * 120)

    # Initialize function
    [fX, fx, dx, d2x, d3x] = init_func(func_name)
    [XMIN, XMAX, YMIN, YMAX, x_min] = init_params(func_name)

    # Set starting point
    if x0 is None:
        if func_name == "Slalom":
            x0 = np.array([[0.5], [0.0]])
        elif func_name == "HairpinTurn":
            x0 = np.array([[0.5], [0.0]])
        else:
            x0 = x_min[0].reshape(-1, 1) if len(x_min) > 0 else np.array([[0.5], [0.0]])

    print(f"\nStarting point: x0 = [{x0[0,0]:.3f}, {x0[1,0]:.3f}]")
    print(f"Function bounds: x ∈ [{XMIN}, {XMAX}], y ∈ [{YMIN}, {YMAX}]")

    # Algorithm parameters
    param_dict = {
        "almton": [0.1, 0.01, 0.1, 2.0],
        "almton_heuristic": [0.1, 0.01, 0.1, 2.0],
        "almton_interp": [
            0.01,
            0.95,
            2.0,
            1e-8,
            0.1,
        ],  # [eta1, eta2, gamma, sigma_min, beta]
        "ar2_interp": [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8],
        "ar3_interp": [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8],
    }

    # Run algorithms and collect paths
    paths_dict = {}
    results_dict = {}

    algorithms = [
        "ALMTON",
        "almton_heuristic",
        "ALMTON-Interp",
        "AR2-Interp",
        "AR3-Interp",
        "Newton",
    ]

    for algo_name in algorithms:
        print(f"\nRunning {algo_name}...")
        print("-" * 120)

        try:
            if algo_name == "ALMTON":
                result, path = run_almton_with_path(
                    fx,
                    dx,
                    d2x,
                    d3x,
                    x0,
                    max_iterations,
                    tol,
                    param_dict["almton"],
                    sdp_tol=sdp_tol,
                )
            elif algo_name == "almton_heuristic":
                result, path = run_almton_heuristic_with_path(
                    fx,
                    dx,
                    d2x,
                    d3x,
                    x0,
                    max_iterations,
                    tol,
                    param_dict["almton_heuristic"],
                    sdp_tol=sdp_tol,
                )
            elif algo_name == "ALMTON-Interp":
                result, path = run_almton_interp_with_path(
                    fx,
                    dx,
                    d2x,
                    d3x,
                    x0,
                    max_iterations,
                    tol,
                    param_dict["almton_interp"],
                    sdp_tol=sdp_tol,
                )
            elif algo_name == "AR2-Interp":
                result, path = run_ar2_with_path(
                    fx, dx, d2x, x0, max_iterations, tol, param_dict["ar2_interp"]
                )
            elif algo_name == "AR3-Interp":
                result, path = run_ar3_with_path(
                    fx, dx, d2x, d3x, x0, max_iterations, tol, param_dict["ar3_interp"]
                )
            elif algo_name == "Newton":
                result, path = run_newton_with_path(
                    fx, dx, d2x, d3x, x0, max_iterations, tol
                )

            paths_dict[algo_name] = path
            results_dict[algo_name] = result

            if result.get("converged", False):
                final_x = result.get("x_final", path[-1].reshape(-1, 1))
                final_f = fx(final_x)
                final_grad = LA.norm(dx(final_x))
                print(
                    f"  Converged: {result.get('iterations', len(path))} iterations, "
                    f"f={final_f:.6e}, ||∇f||={final_grad:.6e}"
                )
            else:
                print(
                    f"  Did not converge after {result.get('iterations', len(path))} iterations"
                )

        except Exception as e:
            print(f"  Error: {str(e)}")
            paths_dict[algo_name] = None
            results_dict[algo_name] = {"error": str(e)}

    # Create visualization
    if save_path:
        os.makedirs(save_path, exist_ok=True)
        fig_path = os.path.join(save_path, f"{func_name}_trajectory_comparison.pdf")
    else:
        fig_path = None

    fig = plot_trajectory_comparison(
        func_name, paths_dict, x0, x_min, XMIN, XMAX, YMIN, YMAX, save_path=fig_path
    )

    # Save data
    if save_path:
        data_path = os.path.join(save_path, f"{func_name}_trajectory_data.pkl")
        pickle_data = {
            "func_name": func_name,
            "x0": x0,
            "paths_dict": paths_dict,
            "results_dict": results_dict,
            "x_min": x_min,
            "bounds": [XMIN, XMAX, YMIN, YMAX],
        }
        with open(data_path, "wb") as f:
            pickle.dump(pickle_data, f)
        print(f"\nData saved to: {data_path}")

    # Print summary
    print("\n" + "=" * 120)
    print("Summary")
    print("=" * 120)
    for algo_name in algorithms:
        if algo_name in results_dict and paths_dict.get(algo_name) is not None:
            path = paths_dict[algo_name]
            result = results_dict[algo_name]
            n_iter = result.get("iterations", len(path))
            converged = result.get("converged", False)
            path_length = np.sum(
                [LA.norm(path[i + 1] - path[i]) for i in range(len(path) - 1)]
            )
            print(
                f"{algo_name:15s}: {n_iter:3d} iter, "
                f"converged={converged}, path_length={path_length:.4f}"
            )

    return {
        "paths_dict": paths_dict,
        "results_dict": results_dict,
        "fig": fig,
    }


# ============================================================================
# Experiment 3.2: Dense Grid & Tabular Analysis (Slalom & HairpinTurn)
# ============================================================================


def generate_dense_grid_3_2(func_name, grid_size=(20, 20)):
    """
    Generate dense grid points for Experiment 3.2.

    Parameters:
    - func_name: Function name
    - grid_size: Grid size (nx, ny)

    Returns:
    - list: List of grid points
    """
    [XMIN, XMAX, YMIN, YMAX, x_min] = init_params(func_name)

    x_vals = np.linspace(XMIN, XMAX, grid_size[0])
    y_vals = np.linspace(YMIN, YMAX, grid_size[1])

    grid_points = []
    for x in x_vals:
        for y in y_vals:
            grid_points.append(np.array([[x], [y]]))

    return grid_points, x_vals, y_vals


def run_single_algorithm_grid_3_2(args):
    """
    Run a single algorithm on a single starting point (for parallel computing).

    Parameters:
    - args: (x0, func_name, algorithm_name, max_iterations, tol, param_dict, sdp_tol)

    Returns:
    - dict: Result dictionary
    """
    x0, func_name, algorithm_name, max_iterations, tol, param_dict, sdp_tol = args

    [fX, fx, dx, d2x, d3x] = init_func(func_name)
    x0 = np.array(x0).reshape(-1, 1)

    result = {
        "algorithm": algorithm_name,
        "x0": x0.flatten(),
        "converged": False,
        "iterations": max_iterations,
        "final_f": np.inf,
        "final_grad_norm": np.inf,
        "time_elapsed": 0.0,
        "failure_reason": None,
    }

    start_time = time.time()

    try:
        if algorithm_name == "GD (α=0.01)":
            x_final, converged, iters = gradient_descent(
                fx,
                dx,
                x0,
                alpha_init=0.01,
                max_iter=max_iterations,
                tol=tol,
                use_line_search=True,
            )
            result["iterations"] = iters
            result["converged"] = converged
            if converged:
                result["final_f"] = fx(x_final)
                result["final_grad_norm"] = LA.norm(dx(x_final))
            else:
                result["failure_reason"] = "max_iterations_exceeded"

        elif algorithm_name == "GD (α=0.05)":
            x_final, converged, iters = gradient_descent(
                fx,
                dx,
                x0,
                alpha_init=0.05,
                max_iter=max_iterations,
                tol=tol,
                use_line_search=True,
            )
            result["iterations"] = iters
            result["converged"] = converged
            if converged:
                result["final_f"] = fx(x_final)
                result["final_grad_norm"] = LA.norm(dx(x_final))
            else:
                result["failure_reason"] = "max_iterations_exceeded"

        elif algorithm_name == "Second-Order Newton":
            x_final, converged, iters = newton_run(
                fx, dx, d2x, d3x, x0, max_iterations, tol, use_line_search=True
            )
            result["iterations"] = iters
            result["converged"] = converged
            if converged:
                result["final_f"] = fx(x_final)
                result["final_grad_norm"] = LA.norm(dx(x_final))
            else:
                if iters >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "saddle_point"

        elif algorithm_name == "Unregularized Third-Order Newton":
            x_final, converged, iters = unregularized_third_newton_run(
                fx, dx, d2x, d3x, x0, max_iterations, tol
            )
            result["iterations"] = iters
            result["converged"] = converged
            if converged:
                result["final_f"] = fx(x_final)
                result["final_grad_norm"] = LA.norm(dx(x_final))
            else:
                if iters >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "sdp_subproblem_failure"

        elif algorithm_name == "AR2-Interp":
            ar2_interp_params = param_dict.get(
                "ar2_interp", [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8]
            )
            r = ar2_interp(
                fx, dx, d2x, x0, max_iterations, tol, ar2_interp_params, verbose=False
            )
            result["iterations"] = r["iterations"]
            result["converged"] = r["converged"]
            if r["converged"]:
                result["final_f"] = r["f_history"][-1]
                result["final_grad_norm"] = r["grad_norm_history"][-1]
            else:
                if r["iterations"] >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "subproblem_failure"

        elif algorithm_name == "AR3-Interp":
            ar3_interp_params = param_dict.get(
                "ar3_interp", [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8]
            )
            r = ar3_interp(
                fx,
                dx,
                d2x,
                d3x,
                x0,
                max_iterations,
                tol,
                ar3_interp_params,
                verbose=False,
            )
            result["iterations"] = r["iterations"]
            result["converged"] = r["converged"]
            if r["converged"]:
                result["final_f"] = r["f_history"][-1]
                result["final_grad_norm"] = r["grad_norm_history"][-1]
            else:
                if r["iterations"] >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "subproblem_failure"

        elif algorithm_name == "ALMTON":
            almton_params = param_dict.get("almton", [0.1, 0.01, 0.1, 2.0])
            r = almton_simple(
                fx,
                dx,
                d2x,
                d3x,
                x0,
                max_iterations,
                tol,
                almton_params,
                verbose=False,
                sdp_tol=sdp_tol,
            )
            result["iterations"] = r["iterations"]
            result["converged"] = r["converged"]
            if r["converged"]:
                result["final_f"] = r["f_history"][-1]
                result["final_grad_norm"] = r["grad_norm_history"][-1]
            else:
                if r.get("sigma_exceeded", False):
                    result["failure_reason"] = "sigma_exceeded"
                elif r["iterations"] >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "sdp_solver_failure"

        elif algorithm_name == "almton_heuristic":
            almton_heuristic_params = param_dict.get(
                "almton_heuristic_heuristic", [0.1, 0.01, 0.1, 2.0]
            )
            r = almton_heuristic(
                fx,
                dx,
                d2x,
                d3x,
                x0,
                max_iterations,
                tol,
                almton_heuristic_params,
                verbose=False,
                sdp_tol=sdp_tol,
            )
            result["iterations"] = r["iterations"]
            result["converged"] = r["converged"]
            if r["converged"]:
                result["final_f"] = r["f_history"][-1]
                result["final_grad_norm"] = r["grad_norm_history"][-1]
            else:
                if r.get("sigma_exceeded", False):
                    result["failure_reason"] = "sigma_exceeded"
                elif r["iterations"] >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "sdp_solver_failure"

        elif algorithm_name == "ALMTON-Interp":
            almton_interp_params = param_dict.get(
                "almton_interp", [0.01, 0.95, 2.0, 1e-8, 0.1]
            )
            r = almton_interp(
                fx,
                dx,
                d2x,
                d3x,
                x0,
                max_iterations,
                tol,
                almton_interp_params,
                verbose=False,
                sdp_solver="auto",
            )
            result["iterations"] = r["iterations"]
            result["converged"] = r["converged"]
            if r["converged"]:
                result["final_f"] = r["f_history"][-1]
                result["final_grad_norm"] = r["grad_norm_history"][-1]
            else:
                if r.get("sigma_exceeded", False):
                    result["failure_reason"] = "sigma_exceeded"
                elif r["iterations"] >= max_iterations:
                    result["failure_reason"] = "max_iterations_exceeded"
                else:
                    result["failure_reason"] = "sdp_solver_failure"

        result["total_time"] = time.time() - start_time

    except Exception as e:
        result["error"] = str(e)
        result["converged"] = False
        result["failure_reason"] = "exception"

    return result


def experiment_3_2_1(
    func_name,
    grid_size=(20, 20),
    max_iterations=1000,
    tol=1e-6,
    use_parallel=True,
    n_jobs=None,
    save_path=None,
    sdp_tol=1e-8,
):
    """
    Experiment 3.2.1: Dense Grid Test for Slalom and HairpinTurn

    Parameters:
    - func_name: Function name ('Slalom' or 'HairpinTurn')
    - grid_size: Grid size
    - max_iterations: Maximum number of iterations (default: 1000)
    - tol: Convergence tolerance (default: 1e-6)
    - use_parallel: Whether to use parallel computing
    - n_jobs: Number of parallel processes
    - save_path: Path to save results
    - sdp_tol: SDP tolerance for ALMTON algorithms (default: 1e-8)

    Returns:
    - dict: Result dictionary
    """
    print("=" * 120)
    print(f"Experiment 3.2.1: Dense Grid Test - {func_name}")
    print("=" * 120)

    # Generate grid points
    grid_points, x_vals, y_vals = generate_dense_grid_3_2(func_name, grid_size)
    print(f"Generated {len(grid_points)} grid points ({grid_size[0]}x{grid_size[1]})")

    # Algorithm list (including ALMTON-Interp)
    algorithms = [
        "GD (α=0.01)",
        "GD (α=0.05)",
        "Second-Order Newton",
        "Unregularized Third-Order Newton",
        "AR2-Interp",
        "AR3-Interp",
        "ALMTON",
        "almton_heuristic",
        "ALMTON-Interp",
    ]

    # Parameter settings
    param_dict = {
        "ar2_interp": [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8],
        "ar3_interp": [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8],
        "almton": [0.1, 0.01, 0.1, 2.0],
        "almton_heuristic": [0.1, 0.01, 0.1, 2.0],
        "almton_interp": [0.01, 0.95, 2.0, 1e-8, 0.1],
    }

    # Determine number of parallel processes
    if n_jobs is None:
        n_jobs = cpu_count()

    # Run all algorithms
    all_results = {}

    for algo_name in algorithms:
        print(f"\nTesting algorithm: {algo_name}")
        print("-" * 120)

        # Prepare arguments
        args_list = [
            (x0, func_name, algo_name, max_iterations, tol, param_dict, sdp_tol)
            for x0 in grid_points
        ]

        # Run
        if use_parallel and len(grid_points) > 1:
            with Pool(processes=n_jobs) as pool:
                results = pool.map(run_single_algorithm_grid_3_2, args_list)
        else:
            results = []
            for i, args in enumerate(args_list):
                if i % 50 == 0:
                    print(f"  Progress: {i}/{len(args_list)}")
                results.append(run_single_algorithm_grid_3_2(args))

        all_results[algo_name] = results
        print(f"  Completed: {algo_name}")

    # Calculate statistics
    print("\n" + "=" * 120)
    print("Statistics")
    print("=" * 120)

    stats = {}
    for algo_name in algorithms:
        results = all_results[algo_name]

        # Only count successfully converged points
        converged_results = [r for r in results if r["converged"]]
        n_converged = len(converged_results)
        n_total = len(results)
        success_rate = n_converged / n_total if n_total > 0 else 0

        if n_converged > 0:
            iterations = [r["iterations"] for r in converged_results]
            times = [r["total_time"] for r in converged_results]
            sdp_times = [
                r.get("sdp_time", 0.0) for r in converged_results if "sdp_time" in r
            ]

            median_iters = np.median(iterations)
            avg_sdp_time = np.mean(sdp_times) if len(sdp_times) > 0 else 0.0
        else:
            median_iters = np.nan
            avg_sdp_time = np.nan

        # Failure reason statistics
        failure_reasons = {}
        failed_results = [r for r in results if not r["converged"]]
        for r in failed_results:
            reason = r.get("failure_reason", "unknown")
            failure_reasons[reason] = failure_reasons.get(reason, 0) + 1

        stats[algo_name] = {
            "success_rate": success_rate,
            "median_iters": median_iters,
            "avg_sdp_time": avg_sdp_time,
            "n_converged": n_converged,
            "n_total": n_total,
            "failure_reasons": failure_reasons,
        }

        print(
            f"{algo_name:<30} Success Rate: {success_rate:>6.1%}, "
            f"Median Iters: {median_iters:>6.1f}, Avg SDP Time: {avg_sdp_time:>8.4f}s"
        )
        if failure_reasons:
            print(f"  Failure Reasons: {failure_reasons}")

    # Generate heatmaps
    if save_path:
        os.makedirs(save_path, exist_ok=True)
        plot_convergence_heatmaps_experiment_3_2_1(
            all_results, func_name, grid_size, x_vals, y_vals, save_path
        )

    # Save detailed data
    if save_path:
        pickle_path = os.path.join(save_path, f"{func_name}_experiment_3_2_1_data.pkl")
        pickle_data = {
            "all_results": all_results,
            "stats": stats,
            "grid_points": grid_points,
            "grid_size": grid_size,
            "x_vals": x_vals,
            "y_vals": y_vals,
            "func_name": func_name,
            "max_iterations": max_iterations,
            "tol": tol,
        }
        with open(pickle_path, "wb") as f:
            pickle.dump(pickle_data, f)
        print(f"\nDetailed data saved to: {pickle_path}")

    return {
        "all_results": all_results,
        "stats": stats,
        "grid_points": grid_points,
    }


def plot_convergence_heatmaps_experiment_3_2_1(
    all_results, func_name, grid_size, x_vals, y_vals, save_path
):
    """
    Generate convergence heatmaps for Experiment 3.2.1 (3x3 layout: 3 rows, 3 columns, 9 algorithms)
    """
    algorithms = list(all_results.keys())
    nx, ny = grid_size

    # Create 3x3 subplot layout (9 algorithms)
    fig, axes = plt.subplots(3, 3, figsize=(36, 36))
    axes = axes.flatten()

    [XMIN, XMAX, YMIN, YMAX, x_min] = init_params(func_name)

    for idx, algo_name in enumerate(algorithms):
        if idx >= 9:
            break

        ax = axes[idx]
        results = all_results[algo_name]

        # Create grid data
        convergence_grid = np.zeros((nx, ny))
        iteration_grid = np.full((nx, ny), np.nan)

        for i, result in enumerate(results):
            row = i // ny
            col = i % ny
            if result["converged"]:
                convergence_grid[row, col] = 1
                iteration_grid[row, col] = result["iterations"]
            else:
                convergence_grid[row, col] = 0

        # Create heatmap data
        heatmap_data = iteration_grid.copy()
        heatmap_data[np.isnan(heatmap_data)] = -1  # Set unconverged points to -1

        # Create colormap
        colors = ["red"] + plt.cm.PuBu(np.linspace(0, 1, 100)).tolist()
        cmap = ListedColormap(colors)
        vmin, vmax = 0, 1000  # Adjusted for max_iterations=1000
        bounds = [-1.5, -0.5] + list(np.linspace(vmin, vmax, 100))
        norm = BoundaryNorm(bounds, cmap.N)

        # Draw heatmap
        im = ax.imshow(
            heatmap_data,
            cmap=cmap,
            norm=norm,
            extent=[x_vals.min(), x_vals.max(), y_vals.min(), y_vals.max()],
            origin="lower",
            aspect="auto",
        )

        ax.set_title(f"{algo_name}", fontsize=14, fontweight="bold")
        ax.set_xlabel(r"$x_1$", fontsize=12)
        ax.set_ylabel(r"$x_2$", fontsize=12)
        ax.grid(True, alpha=0.3)

    # Hide extra subplots
    for idx in range(len(algorithms), 9):
        axes[idx].set_visible(False)

    # Add shared colorbar
    cbar_ax = fig.add_axes([0.93, 0.15, 0.015, 0.7])
    colors = ["red"] + plt.cm.PuBu(np.linspace(0, 1, 100)).tolist()
    cmap = ListedColormap(colors)
    bounds = [-1.5, -0.5] + list(np.linspace(0, 1000, 100))
    norm = BoundaryNorm(bounds, cmap.N)
    cbar = plt.colorbar(
        plt.cm.ScalarMappable(norm=norm, cmap=cmap),
        cax=cbar_ax,
        boundaries=bounds,
    )
    cbar.set_label("Convergence Iterations", fontsize=12)

    plt.tight_layout(rect=[0, 0, 0.92, 1])

    # Save
    file_path = os.path.join(save_path, f"{func_name}_experiment_3_2_1_heatmap.pdf")
    plt.savefig(file_path, bbox_inches="tight", dpi=300)
    print(f"\nHeatmap saved to: {file_path}")
    plt.close()


def experiment_3_2_2(
    func_name,
    n_trials=None,
    max_iterations=1000,
    tol=1e-6,
    use_parallel=True,
    n_jobs=None,
    save_path=None,
    use_grid_data=True,
    grid_data_path=None,
    sdp_tol=1e-8,
):
    """
    Experiment 3.2.2: Tabular Testing for Slalom and HairpinTurn

    Parameters:
    - func_name: Function name ('Slalom' or 'HairpinTurn')
    - n_trials: Number of random trials (if not using grid data)
    - max_iterations: Maximum iterations (default: 1000)
    - tol: Convergence tolerance (default: 1e-6)
    - use_parallel: Whether to use parallel computing
    - n_jobs: Number of parallel processes
    - save_path: Path to save results
    - use_grid_data: Whether to use grid data from Experiment 3.2.1
    - grid_data_path: Path to grid data pickle file
    - sdp_tol: SDP tolerance for ALMTON algorithms (default: 1e-8)

    Returns:
    - dict: Result dictionary
    """
    print("=" * 120)
    print(f"Experiment 3.2.2: Tabular Testing - {func_name}")
    print("=" * 120)

    # Algorithm list (including ALMTON-Interp)
    algorithms = [
        "GD (α=0.01)",
        "GD (α=0.05)",
        "Second-Order Newton",
        "Unregularized Third-Order Newton",
        "AR2-Interp",
        "AR3-Interp",
        "ALMTON",
        "almton_heuristic",
        "ALMTON-Interp",
    ]

    # Parameter settings
    param_dict = {
        "ar2_interp": [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8],
        "ar3_interp": [0.01, 0.95, 0.5, 0.1, 0.01, 2.0, 1e-8],
        "almton": [0.1, 0.01, 0.1, 2.0],
        "almton_heuristic": [0.1, 0.01, 0.1, 2.0],
        "almton_interp": [0.01, 0.95, 2.0, 1e-8, 0.1],
    }

    # Determine starting points: use grid data from Exp 3.2.1 or random points
    if use_grid_data:
        # Try to load data from Experiment 3.2.1
        if grid_data_path is None:
            # Auto search for pickle file
            if save_path:
                grid_data_path = os.path.join(
                    save_path, f"{func_name}_experiment_3_2_1_data.pkl"
                )
            else:
                grid_data_path = os.path.join(
                    EXPERIMENT_3_RESULTS_DIR, f"{func_name}_experiment_3_2_1_data.pkl"
                )

        if os.path.exists(grid_data_path):
            print(f"\nLoading grid data from Experiment 3.2.1: {grid_data_path}")
            with open(grid_data_path, "rb") as f:
                grid_data = pickle.load(f)
            start_points = grid_data["grid_points"]
            all_results = grid_data["all_results"]
            print(f"  Loaded data for {len(start_points)} grid points")
            print(f"  Using full results from Experiment 3.2.1 for analysis")
        else:
            print(f"\nWarning: Experiment 3.2.1 data file not found: {grid_data_path}")
            print(f"  Will use random initial points (n_trials={n_trials or 10})")
            use_grid_data = False

    if not use_grid_data:
        # Generate random starting points
        [XMIN, XMAX, YMIN, YMAX, x_min] = init_params(func_name)
        np.random.seed(42)
        n_trials = n_trials or 10
        start_points = [
            np.array(
                [
                    [np.random.uniform(XMIN, XMAX)],
                    [np.random.uniform(YMIN, YMAX)],
                ]
            )
            for _ in range(n_trials)
        ]
        all_results = {}

    # Determine number of parallel processes
    if n_jobs is None:
        n_jobs = cpu_count()

    # If using grid data, we already have results, just analyze them
    if not use_grid_data:
        # Run all algorithms
        for algo_name in algorithms:
            print(f"\nTesting algorithm: {algo_name}")
            print("-" * 120)

            # Prepare arguments
            args_list = [
                (x0, func_name, algo_name, max_iterations, tol, param_dict, sdp_tol)
                for x0 in start_points
            ]

            # Run
            if use_parallel and len(start_points) > 1:
                with Pool(processes=n_jobs) as pool:
                    results = pool.map(run_single_algorithm_grid_3_2, args_list)
            else:
                results = []
                for args in args_list:
                    results.append(run_single_algorithm_grid_3_2(args))

            all_results[algo_name] = results
            print(f"  Completed: {algo_name}")

    # Compute statistics (only converged points)
    print("\n" + "=" * 120)
    print("Statistical Results (only converged points)")
    print("=" * 120)

    stats = {}
    detailed_data = []

    for algo_name in algorithms:
        if algo_name not in all_results:
            continue

        results = all_results[algo_name]

        # Only count successfully converged points
        converged_results = [r for r in results if r["converged"]]
        n_converged = len(converged_results)
        n_total = len(results)
        conv_rate = n_converged / n_total if n_total > 0 else 0

        if n_converged > 0:
            iterations = [r["iterations"] for r in converged_results]
            times = [r["total_time"] for r in converged_results]
            final_f = [r["final_f"] for r in converged_results]

            median_iters = np.median(iterations)
            q1_iters = np.percentile(iterations, 25)
            q3_iters = np.percentile(iterations, 75)
            iqr_iters = q3_iters - q1_iters

            median_time = np.median(times)
            q1_time = np.percentile(times, 25)
            q3_time = np.percentile(times, 75)
            iqr_time = q3_time - q1_time

            median_final_f = np.median(final_f)
        else:
            median_iters = np.nan
            q1_iters = np.nan
            q3_iters = np.nan
            iqr_iters = np.nan
            median_time = np.nan
            q1_time = np.nan
            q3_time = np.nan
            iqr_time = np.nan
            median_final_f = np.nan

        # Failure reason statistics
        failure_reasons = {}
        failed_results = [r for r in results if not r["converged"]]
        for r in failed_results:
            reason = r.get("failure_reason", "unknown")
            failure_reasons[reason] = failure_reasons.get(reason, 0) + 1

        stats[algo_name] = {
            "median_iters": median_iters,
            "q1_iters": q1_iters,
            "q3_iters": q3_iters,
            "iqr_iters": iqr_iters,
            "median_time": median_time,
            "q1_time": q1_time,
            "q3_time": q3_time,
            "iqr_time": iqr_time,
            "conv_rate": conv_rate,
            "median_final_f": median_final_f,
            "n_converged": n_converged,
            "n_total": n_total,
            "failure_reasons": failure_reasons,
        }

        # Detailed data
        for result in results:
            detailed_data.append(
                {
                    "Algorithm": algo_name,
                    "x0": str(result["x0"]),
                    "Converged": result["converged"],
                    "Iterations": (
                        result["iterations"] if result["converged"] else np.nan
                    ),
                    "Final f(x)": (
                        result["final_f"] if result["converged"] else np.nan
                    ),
                    "Final ||∇f||": (
                        result["final_grad_norm"] if result["converged"] else np.nan
                    ),
                    "Time (s)": result["total_time"],
                    "Failure Reason": result.get("failure_reason", ""),
                }
            )

    # Print statistics table
    print(
        f"\n{'Algorithm':<30} {'Median Iters (IQR)':<25} {'Median Time(s) (IQR)':<25} {'Conv Rate':<12} {'Median Final f':<15}"
    )
    print("-" * 120)

    for algo_name in algorithms:
        if algo_name not in stats:
            continue
        s = stats[algo_name]
        if not np.isnan(s["median_iters"]):
            iters_str = (
                f"{s['median_iters']:.2f} ({s['q1_iters']:.2f}-{s['q3_iters']:.2f})"
            )
            time_str = f"{s['median_time']:.4f} ({s['q1_time']:.4f}-{s['q3_time']:.4f})"
            print(
                f"{algo_name:<30} {iters_str:<25} {time_str:<25} "
                f"{s['conv_rate']:>6.1%}       {s['median_final_f']:>12.6e}"
            )
        else:
            print(
                f"{algo_name:<30} {'N/A':<25} {'N/A':<25} {s['conv_rate']:>6.1%}       {'N/A':<15}"
            )
        # Print failure reason statistics
        if s["failure_reasons"]:
            print(f"  Failure Reasons:")
            for reason, count in sorted(
                s["failure_reasons"].items(), key=lambda x: -x[1]
            ):
                print(
                    f"    {reason}: {count} ({count/(s['n_total']-s['n_converged'])*100:.1f}%)"
                )

    # Save to Excel
    if save_path:
        os.makedirs(save_path, exist_ok=True)

        # Create DataFrame
        df_detailed = pd.DataFrame(detailed_data)

        # Create summary table
        summary_rows = []
        for algo_name in algorithms:
            if algo_name not in stats:
                continue
            s = stats[algo_name]
            row = {
                "Algorithm": algo_name,
                "Median Iters": s["median_iters"],
                "Q1 Iters": s["q1_iters"],
                "Q3 Iters": s["q3_iters"],
                "IQR Iters": s["iqr_iters"],
                "Median Time(s)": s["median_time"],
                "Q1 Time(s)": s["q1_time"],
                "Q3 Time(s)": s["q3_time"],
                "IQR Time(s)": s["iqr_time"],
                "Conv Rate": s["conv_rate"],
                "Median Final f": s["median_final_f"],
                "N Converged": s["n_converged"],
                "N Total": s["n_total"],
            }
            # Add failure reason statistics
            if s["failure_reasons"]:
                failure_str = "; ".join(
                    [
                        f"{k}:{v}"
                        for k, v in sorted(
                            s["failure_reasons"].items(), key=lambda x: -x[1]
                        )
                    ]
                )
                row["Failure Reasons"] = failure_str
            else:
                row["Failure Reasons"] = ""
            summary_rows.append(row)

        df_summary = pd.DataFrame(summary_rows)

        # Format numeric columns to 6 decimal places
        numeric_cols_summary = [
            "Median Iters",
            "Q1 Iters",
            "Q3 Iters",
            "IQR Iters",
            "Median Time(s)",
            "Q1 Time(s)",
            "Q3 Time(s)",
            "IQR Time(s)",
            "Conv Rate",
            "Median Final f",
        ]
        for col in numeric_cols_summary:
            if col in df_summary.columns:
                df_summary[col] = pd.to_numeric(df_summary[col], errors="coerce").round(
                    6
                )

        numeric_cols_detailed = ["Iterations", "Final f(x)", "Final ||∇f||", "Time (s)"]
        for col in numeric_cols_detailed:
            if col in df_detailed.columns:
                df_detailed[col] = pd.to_numeric(
                    df_detailed[col], errors="coerce"
                ).round(6)

        # Save
        excel_path = os.path.join(save_path, f"{func_name}_experiment_3_2_2.xlsx")
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            df_summary.to_excel(writer, sheet_name="Summary", index=False)
            df_detailed.to_excel(writer, sheet_name="Detailed", index=False)

        # Apply number format to Excel cells
        wb = load_workbook(excel_path)
        for sheet_name in ["Summary", "Detailed"]:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.000000"
        wb.save(excel_path)

        print(f"\nResults saved to: {excel_path}")

        # Save pickle
        pickle_path = os.path.join(save_path, f"{func_name}_experiment_3_2_2_data.pkl")
        pickle_data = {
            "all_results": all_results,
            "stats": stats,
            "detailed_data": detailed_data,
            "func_name": func_name,
            "max_iterations": max_iterations,
            "tol": tol,
        }
        with open(pickle_path, "wb") as f:
            pickle.dump(pickle_data, f)
        print(f"Detailed data saved to: {pickle_path}")

    return {
        "all_results": all_results,
        "stats": stats,
        "detailed_data": detailed_data,
    }


def experiment_3_2(
    func_name="Slalom",
    grid_size=(20, 20),
    max_iterations=1000,
    tol=1e-6,
    use_parallel=True,
    n_jobs=None,
    save_path=None,
    sdp_tol=1e-8,
    n_trials=10,
):
    """
    Experiment 3.2 consolidates the dense grid (3.2.1) and
    tabular (3.2.2) analyses with a default 20x20 grid.
    """
    print("=" * 120)
    print(f"Experiment 3.2: Dense Grid + Tabular Analysis - {func_name}")
    print("=" * 120)

    if save_path is None:
        save_path = EXPERIMENT_3_RESULTS_DIR
    _ensure_dir(save_path)

    dense_result = experiment_3_2_1(
        func_name=func_name,
        grid_size=grid_size,
        max_iterations=max_iterations,
        tol=tol,
        use_parallel=use_parallel,
        n_jobs=n_jobs,
        save_path=save_path,
        sdp_tol=sdp_tol,
    )

    tabular_result = experiment_3_2_2(
        func_name=func_name,
        n_trials=n_trials,
        max_iterations=max_iterations,
        tol=tol,
        use_parallel=use_parallel,
        n_jobs=n_jobs,
        save_path=save_path,
        use_grid_data=True,
        grid_data_path=None,
        sdp_tol=sdp_tol,
    )

    return {"dense": dense_result, "tabular": tabular_result}


# ============================================================================
# Main Function
# ============================================================================

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Run Experiment 3: Trajectory (3.1) or Dense Grid (3.2)"
    )
    parser.add_argument(
        "--experiment",
        type=str,
        default="3.1",
        choices=["3.1", "3.2", "both"],
        help="Choose 3.1 (trajectory), 3.2 (dense grid/tabular), or both",
    )
    parser.add_argument(
        "--function",
        type=str,
        default="Slalom",
        choices=["Slalom", "HairpinTurn"],
        help="Function name (default: Slalom)",
    )
    parser.add_argument(
        "--x0",
        type=float,
        nargs=2,
        default=None,
        help="Starting point [x, y] (default: from function defaults)",
    )
    parser.add_argument(
        "--max_iterations",
        type=int,
        default=1000,
        help="Maximum iterations (default: 1000)",
    )
    parser.add_argument(
        "--tol",
        type=float,
        default=1e-8,
        help="Convergence tolerance (default: 1e-8)",
    )
    parser.add_argument(
        "--save_path",
        type=str,
        default=None,
        help="Save path (default: results/experiment_3/)",
    )
    parser.add_argument(
        "--sdp_tol",
        type=float,
        default=1e-8,
        help="SDP tolerance for ALMTON (default: 1e-8)",
    )
    parser.add_argument(
        "--grid_size_3_2",
        type=int,
        nargs=2,
        default=[20, 20],
        help="Grid size for Experiment 3.2 (default: 20 20)",
    )
    parser.add_argument(
        "--n_trials_3_2",
        type=int,
        default=10,
        help="Random trials for Experiment 3.2 fallback (default: 10)",
    )
    parser.add_argument(
        "--no_parallel",
        action="store_true",
        help="Disable parallel processing for Experiment 3.2",
    )
    parser.add_argument(
        "--n_jobs",
        type=int,
        default=None,
        help="Number of parallel processes for Experiment 3.2",
    )

    args = parser.parse_args()

    # Set save path
    if args.save_path is None:
        args.save_path = EXPERIMENT_3_RESULTS_DIR

    if args.experiment in ["3.1", "both"]:
        # Set starting point for trajectory comparison
        x0 = None
        if args.x0 is not None:
            x0 = np.array([[args.x0[0]], [args.x0[1]]])

        experiment_3_trajectory_comparison(
            func_name=args.function,
            x0=x0,
            max_iterations=args.max_iterations,
            tol=args.tol,
            save_path=args.save_path,
            sdp_tol=args.sdp_tol,
        )

    if args.experiment in ["3.2", "both"]:
        experiment_3_2(
            func_name=args.function,
            grid_size=tuple(args.grid_size_3_2),
            max_iterations=args.max_iterations,
            tol=args.tol,
            use_parallel=not args.no_parallel,
            n_jobs=args.n_jobs,
            save_path=args.save_path,
            sdp_tol=args.sdp_tol,
            n_trials=args.n_trials_3_2,
        )

    print("\n" + "=" * 120)
    print("Experiment 3 execution finished!")
    print("=" * 120)
